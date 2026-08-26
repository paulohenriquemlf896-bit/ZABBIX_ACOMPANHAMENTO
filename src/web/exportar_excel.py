#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera o relatorio de exportacao em Excel (.xlsx) — ver
specs/exportacao_relatorio.md.

Uma aba por periodo selecionado, mais uma aba de capa ("Resumo") com o
que foi filtrado. Cor de severidade sempre a paleta oficial do Zabbix
(mesma de src/relatorios_service.py — nunca uma paleta inventada, ver
prompts/tarefas/frontend.txt item 5).

Dependencia: openpyxl (ver requirements.txt e specs/exportacao_relatorio.md
para a justificativa — biblioteca padrao nao gera .xlsx).
"""

import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from relatorios_service import SEV_NOME, SEV_COR, fmt_ts  # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

CABECALHO_FILL = PatternFill(start_color="F0F2F5", end_color="F0F2F5", fill_type="solid")
CABECALHO_FONT = Font(bold=True)
COLUNAS = ["#", "Problema", "Ocorrencias", "%", "Gravidade", "Hosts afetados", "Ultima vez"]


def _escrever_capa(wb, hosts_selecionados, periodos_titulos, gerado_em):
    aba = wb.active
    aba.title = "Resumo"
    aba.append(["Relatorio de problemas recorrentes - Acompanhamento Zabbix"])
    aba["A1"].font = Font(bold=True, size=14)
    aba.append([])
    aba.append(["Gerado em", gerado_em])
    aba.append(["Periodos", ", ".join(periodos_titulos)])
    aba.append(["Hosts", ", ".join(hosts_selecionados) if hosts_selecionados else "Todos"])
    aba.column_dimensions["A"].width = 16
    aba.column_dimensions["B"].width = 70


def _escrever_periodo(wb, dados):
    aba = wb.create_sheet(title=dados["titulo"][:31])
    aba.append(COLUNAS)
    for celula in aba[1]:
        celula.fill = CABECALHO_FILL
        celula.font = CABECALHO_FONT

    if not dados["ranking"]:
        aba.append(["Nenhum problema registrado neste periodo."])
    else:
        total = dados["total"]
        for i, g in enumerate(dados["ranking"], 1):
            pct = round(g["count"] / total * 100, 1) if total else 0
            linha = [
                i, g["nome"], g["count"], pct,
                SEV_NOME.get(g["sev"], "?"),
                ", ".join(sorted(g["hosts"])),
                fmt_ts(g["ultimo"]),
            ]
            aba.append(linha)
            cor = SEV_COR.get(g["sev"], "#97AAB3").lstrip("#")
            aba.cell(row=aba.max_row, column=5).fill = PatternFill(
                start_color=cor, end_color=cor, fill_type="solid"
            )

    aba.freeze_panes = "A2"
    larguras = [4, 55, 12, 8, 12, 40, 18]
    for col, largura in zip("ABCDEFG", larguras):
        aba.column_dimensions[col].width = largura
    for celula in aba["B"]:
        celula.alignment = Alignment(wrap_text=False, vertical="top")


def gerar_excel(dados_por_periodo: dict, hosts_selecionados: list, gerado_em: str) -> bytes:
    """dados_por_periodo: {periodo: {"titulo","desde","ranking","total",
    "por_severidade"}} — mesmo shape de
    src/web/services/exportacao.py:dados_exportacao(). Devolve os bytes
    do arquivo .xlsx pronto para download."""
    wb = Workbook()
    titulos = [d["titulo"] for d in dados_por_periodo.values()]
    _escrever_capa(wb, hosts_selecionados, titulos, gerado_em)
    for dados in dados_por_periodo.values():
        _escrever_periodo(wb, dados)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
